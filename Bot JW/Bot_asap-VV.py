from functools import partial
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import StaleElementReferenceException, WebDriverException
import getpass
import easygui
from datetime import datetime
import sys
import pyautogui as gui
import time
import math
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException

class Dados:

    def __init__(self,data):
        self.data=data
    def tratamento_de_usuario(self):
        op = input('Tentar novamente (s/n)?: ')
        if op in ['s','n','S','N']:
            if op == 's' or op == 'S':
                print()
                usuario, senha = self.dados_usuario()
                return usuario , senha
            if op == 'n' or op == 'N':
                print('Programa Finalizado!')
                self.driver.quit()
                sys.exit()
        else:
            print('Opção invalida:')
            print('Programa Finalizado!')
            self.driver.quit()
            sys.exit()
            

    def dados_usuario(self):
        login = 'leticia.almeida@viavarejo.com.br' #input('Insira email de acesso: ')
        password = 'Vvlog102030' #getpass.getpass('Password (Texto Oculto): ')

        if login == '' or password == '':
            print('Preencha todos os campos corretamente!!')
            self.tratamento_de_usuario()
        else:    
            return login, password

    def valida_login(self,usuario, senha):
        if usuario == '' or  senha == '':
            print('Existem campos Vazios!!')
            self.valida_login(usuario,senha)
        else:
            return usuario, senha
        #    login_asap(self.driver,usuario,senha)
    
    def login_email(self): #,usuario,senha
        url='https://webmail1.hostinger.com.br/'
        self.driver = webdriver.Chrome(executable_path=r'./chromedriver.exe')
        self.driver.get(url)
        op = True
        while op:
            user = self.driver.find_element_by_id('username')
            password = self.driver.find_element_by_id('password')
            btn_login = self.driver.find_element_by_id('submit')
            logado = self.driver.find_element_by_id('remember_me')
            self.usuario, self.senha = self.dados_usuario()
            user.send_keys(self.usuario)
            password.send_keys(self.senha)
            logado.click()
            btn_login.click()
            erro = self.driver.find_elements_by_xpath('//*[@id="login"]/div/div[2]/div[1 and (@class ="alert alert-danger")]')
            if len(erro)>0:
                print(erro[0].text)
                print('')
                self.usuario, self.senha = self.tratamento_de_usuario()
                if self.usuario == '' and self.senha == '':
                    op = False
                    print('Programa Finalizado')
                else:
                    op = True
            else:
                self.roteiro_direto(self.driver)
                self.extrair_dados_tabela()
                self.driver.quit()
                sys.exit()
      
    def roteiro_direto(self,driver):
        menu = driver.find_element_by_xpath('//*[@id="menu-monitorar"]/a')
        roteiro_direto = driver.find_element_by_xpath('//*[@id="menu-monitorar"]/ul/li[6]/a')
        menu.click()
        roteiro_direto.click()
        filtro = driver.find_element_by_xpath('//*[@id="filtroButton"]')
        selecionar = driver.find_element_by_xpath('//*[@id="filtroSelect"]')
        transit_point = driver.find_element_by_xpath('//*[@id="filtroSelect"]/option[16]')
        btn_filtro = driver.find_element_by_xpath('//*[@id="filterApply"]')
        contendo = driver.find_element_by_xpath('//*[@id="filterInputs"]/div/div[1]')
        filtro.click()
        selecionar.click()
        transit_point.click()
        contendo.click()
        gui.write('Sim')
        time.sleep(3)
        gui.press('down')
        gui.press('Enter')
        btn_filtro.click()
        filtro_data = driver.find_element_by_xpath('//*[@id="buttonDateRange"]')
        filtro_data.click()
        filtro_dia = driver.find_element_by_xpath('/html/body/div[5]/div[3]/ul/li[7]')
        filtro_dia.click()
        self.driver.find_element_by_xpath('/html/body/div[5]/div[1]/div[1]/input').click()
        gui.hotkey('ctrl','a')
        gui.write(self.data)
        self.driver.find_element_by_xpath('/html/body/div[5]/div[2]/div[1]/input').click()
        gui.hotkey('ctrl','a')
        gui.write(self.data)
        filtro_data.click()


    def qtde_paginas(self):
        sleep(5)
        pagina = self.driver.find_element_by_xpath('//div[@class="registros text-muted"]')
        #pagina = self.driver.find_element_by_xpath('//*[@id="main-content"]/div/table/tfoot/tr/td/div[2]')
        pagina = str(pagina.text)
        num_pagina = str(pagina.split("/")[0])
        num_pagina1 = str(num_pagina.split(" ")[0])
        num_pagina4 = int(num_pagina1.split("–")[1])
        num_pagina3 = int(num_pagina.split(" ")[2])
        num_considerar = math.ceil(num_pagina3/num_pagina4)
        if num_considerar == 0:
            print('não existe conteudo')
            self.driver.quit()
        else:
            num_considerar
            return  num_considerar

    def qtde_linhas(self):
        sleep(2)
        linhas = self.driver.find_elements_by_xpath('//table/tbody/tr')
        num_lin = len(linhas)
        return num_lin

    def extrair_dados_tabela(self):
        wdw = WebDriverWait(self.driver,15)
        CRIADO_EM = []
        STATUS = []
        CODIGO_INTERNO = []
        NUM_NFE = []
        NUM_CTE = []
        ATRASO = []
        TRANSFERENCIA = []
        CATEGORIA_PEDIDO = []
        CLIENTE = []
        ORIGEM = []
        DESTINO = []

        pg = self.qtde_paginas()
        # pg = 1
        for pagina in range(1,(pg + 1)):
            lin = self.qtde_linhas()
            print(f'extraindo dados: {pagina} de {pg} pag | linhas {lin} ...')
            for l in range(1,(lin + 1)):
                for c in range(1,12):
                    valor_campo = '//table/tbody/tr['+ str(l) + ']/td['+ str(c) + ']'

                    try:
                        element = self.driver.find_element_by_xpath(valor_campo)
                    except (StaleElementReferenceException, WebDriverException) as e:
                        espera = partial(self.esperar_elemento,By.XPATH, '//*[@id="hojeDireto"]/div/div[1]/div[1]/button')
                        valida = wdw.until(espera)
                        if valida == True:
                            element = self.driver.find_element_by_xpath(valor_campo)
                            # print(f'erro na pagina {pg} linha {l}')
                            continue
                        else:
                            print(f'erro na pagina {pg} linha {l}')
                            sys.exit()

                    if c == 1:
                        CRIADO_EM.append(element.text)
                    elif c == 2:
                        STATUS.append(element.text)
                    elif c == 3:
                        CODIGO_INTERNO.append(element.text) 
                    elif c == 4:
                        NUM_NFE.append(element.text)
                    elif c == 5:
                        NUM_CTE.append(element.text)
                    elif c== 6:
                        ATRASO.append(element.text)
                    elif c== 7:
                        TRANSFERENCIA.append(element.text)
                    elif c== 8:
                        CATEGORIA_PEDIDO.append(element.text)
                    elif c== 9:
                        CLIENTE.append(element.text)
                    elif c== 10:
                        ORIGEM.append(element.text)
                    elif c== 11:
                        DESTINO.append(element.text)

                    else:
                        continue
            sleep(2)
            if self.valida_elemento(By.CLASS_NAME, 'nextLink') == True:
                espera_btn = partial(self.esperar_elemento,By.CLASS_NAME, 'nextLink')
                valida_btn = wdw.until(espera_btn)
                if valida_btn == True:
                    btn_prox = self.driver.find_element_by_class_name('nextLink')
                    btn_prox.click()
            else:
                self.driver.quit()
                #sys.exit()

        op2 = input('Gerar arquivo Excel (s/n)?: ')
        if op2 in ['s','n','S','N']:
            if op2 == 's' or op2 == 'S':
                print()
                diretorio = self.caminho_arquivo_excel()
                if diretorio == '':
                    print('Caminho definido não é valido... Dados não salvos!')
                    self.driver.quit()
                    sys.exit()
                else:
                    self.cria_planilha(diretorio,CRIADO_EM,STATUS, CODIGO_INTERNO,NUM_NFE,NUM_CTE,ATRASO,TRANSFERENCIA,CATEGORIA_PEDIDO,CLIENTE, ORIGEM, DESTINO)
                    print('Processo Concluido')
            elif op2 == 'n' or op2 == 'N':
                print('Dados não salvos!')
                self.driver.quit()
                sys.exit()
            else:
                print('Dados não salvos!')
                self.driver.quit()
                sys.exit()
    

    def cria_planilha(self,dirSave,CRIADO_EM,STATUS, CODIGO_INTERNO,NUM_NFE,NUM_CTE,ATRASO,TRANSFERENCIA,CATEGORIA_PEDIDO,CLIENTE, ORIGEM, DESTINO):
        index = 2
        wb = openpyxl.Workbook()
        active_sheet = wb.active  # Pegando sheet ativa
        active_sheet.title = "BASE_DADOS"  # Mudando titulo
        active_sheet['A1']='CRIADO EM'
        active_sheet['B1']='STATUS'
        active_sheet['C1']='CODIGO INTERNO'
        active_sheet['D1']='NUM NFE'
        active_sheet['E1']='NUM CTE'
        active_sheet['F1']='ATRASO'
        active_sheet['G1']='TRANSFERENCIA'
        active_sheet['H1']='CATEGORIA DO PEDIDO'
        active_sheet['I1']='CLIENTE'
        active_sheet['J1']='ORIGEM'
        active_sheet['K1']='DESTINO'
        for criado_em, status, codigo_interno,num_nfe,num_cte,atraso,transferencia,categoria_pedido, cliente,origem, destino in zip(CRIADO_EM,STATUS, CODIGO_INTERNO,NUM_NFE,NUM_CTE,ATRASO,TRANSFERENCIA,CATEGORIA_PEDIDO,CLIENTE, ORIGEM, DESTINO):
            active_sheet.cell(column=1,row=index,value=criado_em)
            active_sheet.cell(column=2,row=index,value=status)
            active_sheet.cell(column=3,row=index,value=codigo_interno)
            active_sheet.cell(column=4,row=index,value=num_nfe)
            active_sheet.cell(column=5,row=index,value=num_cte)
            active_sheet.cell(column=6,row=index,value=atraso)
            active_sheet.cell(column=7,row=index,value=transferencia)
            active_sheet.cell(column=8,row=index,value=categoria_pedido)
            active_sheet.cell(column=9,row=index,value=cliente)
            active_sheet.cell(column=10,row=index,value=origem)
            active_sheet.cell(column=11,row=index,value=destino)
            index += 1
        data_hora = datetime.now()
        data_format = data_hora.strftime('%d%m%Y_%H%M%S')
        wb.save(dirSave + 'Entregas_' + data_format +'.xlsx')

    def caminho_arquivo_excel(self):
        diretorio = easygui.diropenbox()
        return diretorio +'\\'

    def esperar_elemento(self, by, element, driver):
        return bool(driver.find_elements(by,element))

    def valida_elemento(self,tipo, path):

        try: self.driver.find_element(by=tipo,value=path)

        except NoSuchElementException as e: return False

        return True


data = input('digitar data: ')
# try:
extrair_dados = Dados(data)
extrair_dados.login_email()

# except:
#     print('Erro inesperado tente novamente!!')
#     sys.exit()
# 'leticia.daiane@viavarejo.com.br','Socorro@01'

# def esperar_elemento(by, element, driver):
#     return bool(driver.find_elements(by,element))
# url='https://app.asaplog.com.br/'
# driver = webdriver.Chrome(executable_path=r'./chromedriver.exe')
# driver.get(url)
# wdw = WebDriverWait(driver,15)
# # # espera = partial(esperar_elemento,By.XPATH, '//*[@id="login"]/div/div[2]/div[1 and (@class ="alert alert-danger")]')
# # espera = partial(esperar_elemento,By.XPATH, '//*[@id="hojeDireto"]/div/div[1]/div[1]/button')
# # wdw.until_not(espera)
# # d= driver.find_element_by_xpath('//*[@id="qtdHojeDireto"]')
# # assert d.text != '','erro'
# # dd = driver.f
# espera_btn = partial(esperar_elemento,By.XPATH, '//*[@id="RoteirosDia_next"]')
# valida_btn = wdw.until(espera_btn)

